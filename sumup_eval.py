import json
import os
import argparse
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import xlsxwriter 

# 定义所有评估数据集结构
DATASET_STRUCTURE = {
    "TIF(Text Instruction Following)": {
        "subsets": ["continuation", "creation", "empathy", "recommendation", "rewriting", "safety", "simulation"],
        "convert_to_percent": True,
        "score_key": "score"
    },
    "TIF-En(English Text Instruction Following)": {
        "subsets": ["continuation_en", "creation_en", "empathy_en", "recommendation_en", "rewriting_en", "safety_en", "simulation_en"],
        "convert_to_percent": True,
        "score_key": "score"
    },
    "SIF(Speech Instruction Following)": {
        "subsets": ["emotional_control", "language_control", "non_verbal_vocalization", "pacing_control", "style_control", "volume_control"],
        "convert_to_percent": True,
        "score_key": "score"
    },
    "SIF-En(English Speech Instruction Following)": {
        "subsets": ["emotional_control_en", "language_control_en", "non_verbal_vocalization_en", "pacing_control_en", "style_control_en", "volume_control_en"],
        "convert_to_percent": True,
        "score_key": "score"
    },
    "MTD(Multi-turn Dialogue)": {
        "subsets": ["progression", "backtracking", "transition"],
        "convert_to_percent": True,
        "score_key": "score"
    },
    "GK(General Knowledge)": {
        "subsets": ["general_knowledge"],
        "parts": {
            "general_knowledge": ["mathematics", "geography", "politics", "chemistry", "biology", "law", "physics", "history", "medicine", "economics", "sports", "culture"]
        },
        "convert_to_percent": True,
        "score_key": "acc",
        "has_invalid": False,  # 标记GK没有invalid字段
        "is_accuracy": True  # 标记GK是准确率数据集
    },
    "ML(Mathematical Logic)": {
        "subsets": ["basic_math", "math", "logical_reasoning"],
        "parts": {
            "logical_reasoning": ["analysis", "induction", "analogy", "logic"]
        },
        "convert_to_percent": True,
        "score_key": "acc",
        "has_invalid": False,  # 标记ML没有invalid字段
        "is_accuracy": True  # 标记ML是准确率数据集
    },
    "DC(Discourse Comprehension)": {
        "subsets": ["discourse_comprehension"],
        "parts": {
            "discourse_comprehension": ["inference", "induction", "analysis"]
        },
        "convert_to_percent": True,
        "score_key": "acc",
        "has_invalid": False,  # 标记DC没有invalid字段
        "is_accuracy": True  # 标记DC是准确率数据集
    },
    "SV(Speaker Variations)": {
        "subsets": ["age", "accent", "volume", "speed"],
        "control_groups": ["age_cmp", "accent_cmp", "volume_cmp", "speed_cmp"],
        "parts": {
            "age": ["child", "elder"],
            "accent": ["tianjin", "beijing", "dongbei", "sichuan"],
            "volume": ["down", "up"],
            "speed": []
        },
        "convert_to_percent": True,
        "score_key": "score",
        "has_invalid": True,
        "is_accuracy": False
    },
    "EV(Environmental Variations)": {
        "subsets": ["non_vocal_noise", "vocal_noise", "unstable_signal"],
        "control_groups": ["non_vocal_noise_cmp", "vocal_noise_cmp", "unstable_signal_cmp"],
        "parts": {
            "non_vocal_noise": ["echo", "outdoors", "farField"],
            "vocal_noise": ["tvPlayback", "backgroundChat", "vocalMusic", "voiceAnnouncement"],
            "unstable_signal": []
        },
        "convert_to_percent": True,
        "score_key": "score",
        "has_invalid": True,
        "is_accuracy": False
    },
    "CV(Content Variations)": {
        # "subsets": ["fillers", "repetition", "mispronunciation", "grammatical_error", "topic_shift", "code_switching"],
        # "control_groups": ["fillers_cmp", "repetition_cmp", "mispronunciation_cmp", "grammatical_error_cmp", "topic_shift_cmp", "code_switching_cmp"],
        "subsets": ["casual_talk", "mispronunciation", "grammatical_error", "topic_shift", "code_switching"],
        "control_groups": ["casual_talk_cmp", "mispronunciation_cmp", "grammatical_error_cmp", "topic_shift_cmp", "code_switching_cmp"],
        "parts": {
            # "fillers": [],
            # "repetition": [],
            "casual_talk": [],
            "mispronunciation": [],
            "grammatical_error": [],
            "topic_shift": [],
            "code_switching": []
        },
        "convert_to_percent": True,
        "score_key": "score",
        "has_invalid": True,
        "is_accuracy": False
    }
}

def convert_score(score, convert_to_percent, is_accuracy=False):
    """Convert score: if needed and less than 5, convert to percentage"""
    if not convert_to_percent:
        return score
        
    # 对于准确率数据集(GK, ML, DC)，如果分数小于1，乘以100转换为百分比
    if is_accuracy and score < 1:
        return score * 100
    
    # 对于其他数据集，如果分数小于5，乘以20转换为百分比
    if not is_accuracy and score < 5:
        return score * 20
        
    return score

def extract_score(data, score_key):
    """Extract score from data, handling both direct scores and nested dictionaries"""
    if isinstance(data, (int, float)):
        return data
    elif isinstance(data, dict):
        # Try to get score from dictionary
        if score_key in data:
            return data[score_key]
        elif "score" in data:
            return data["score"]
        elif "acc" in data:
            return data["acc"]
    return None

def load_performance_data(model, subset_name, result_dir, wasr=False):
    """Load performance data for specified model and subset"""
    # Check if it's an SIF subset
    # import pdb;pdb.set_trace()
    if subset_name in DATASET_STRUCTURE["SIF(Speech Instruction Following)"]["subsets"] or subset_name in DATASET_STRUCTURE["SIF-En(English Speech Instruction Following)"]["subsets"]:
        file_path = f"{result_dir}/{model}/{subset_name}/{model}_{subset_name}_gpt-4o-audio_performance.json"
    else:
        if not wasr:
            file_path = f"{result_dir}/{model}/{subset_name}/{model}_{subset_name}_default_performance.json"
        else:
            file_path = f"{result_dir}/{model}/{subset_name}/{model}_{subset_name}_wasr_default_performance.json"
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data.get("performance", {})
    except (FileNotFoundError, json.JSONDecodeError):
        print(f"Error loading performance data from {file_path}")
        return None

def calculate_weighted_average(scores_with_weights):
    """Calculate weighted average"""
    if not scores_with_weights:
        return None
    
    total_weighted_score = 0
    total_weight = 0
    
    for score, weight in scores_with_weights:
        total_weighted_score += score * weight
        total_weight += weight
    
    return total_weighted_score / total_weight if total_weight > 0 else None

def calculate_dataset_average(model, dataset_name, dataset_info, result_dir, wasr=False):
    """Calculate weighted average score for dataset"""
    subsets = dataset_info["subsets"]
    convert_to_percent = dataset_info["convert_to_percent"]
    score_key = dataset_info["score_key"]
    has_invalid = dataset_info.get("has_invalid", True)  # 默认为True
    is_accuracy = dataset_info.get("is_accuracy", False)  # 默认为False
    
    scores_with_weights = []
    control_scores_with_weights = []
    all_subsets_exist = True
    
    # Check experimental subsets
    for subset in subsets:
        performance_data = load_performance_data(model, subset, result_dir, wasr)
        if performance_data is None:
            all_subsets_exist = False
            continue
            
        # Get main score and weight (number of valid data points)
        if subset in performance_data:
            subset_data = performance_data[subset]
            score = extract_score(subset_data, score_key)
            if score is not None:
                converted_score = convert_score(score, convert_to_percent, is_accuracy)
                
                # 根据是否有invalid字段计算权重
                if has_invalid:
                    total = subset_data.get("total", 0) if isinstance(subset_data, dict) else 0
                    invalid = subset_data.get("invalid", 0) if isinstance(subset_data, dict) else 0
                    weight = total - invalid  # Number of valid data points as weight
                else:
                    # 对于GK、ML、DC，使用valid字段作为权重
                    weight = subset_data.get("valid", 0) if isinstance(subset_data, dict) else 0
                
                if weight > 0:
                    scores_with_weights.append((converted_score, weight))
                else:
                    print(f"Warning: {subset} has 0 valid data points")
            else:
                all_subsets_exist = False
        else:
            all_subsets_exist = False
    
    # Check control groups (if they exist)
    if "control_groups" in dataset_info:
        for control_group in dataset_info["control_groups"]:
            performance_data = load_performance_data(model, control_group, result_dir, wasr)
            if performance_data is None:
                all_subsets_exist = False
                continue
                
            if control_group in performance_data:
                control_data = performance_data[control_group]
                score = extract_score(control_data, score_key)
                if score is not None:
                    converted_score = convert_score(score, convert_to_percent, is_accuracy)
                    
                    # 根据是否有invalid字段计算权重
                    if has_invalid:
                        total = control_data.get("total", 0) if isinstance(control_data, dict) else 0
                        invalid = control_data.get("invalid", 0) if isinstance(control_data, dict) else 0
                        weight = total - invalid
                    else:
                        # 对于GK、ML、DC，使用valid字段作为权重
                        weight = control_data.get("valid", 0) if isinstance(control_data, dict) else 0
                    
                    if weight > 0:
                        control_scores_with_weights.append((converted_score, weight))
                    else:
                        print(f"Warning: {control_group} has 0 valid data points")
                else:
                    all_subsets_exist = False
            else:
                all_subsets_exist = False
    
    # Calculate weighted average
    avg_score = calculate_weighted_average(scores_with_weights)
    avg_control_score = calculate_weighted_average(control_scores_with_weights)
    
    return all_subsets_exist, avg_score, avg_control_score

def print_subset_results(model, dataset_name, dataset_info, result_dir, wasr=False):
    """Print detailed results for subsets"""
    subsets = dataset_info["subsets"]
    convert_to_percent = dataset_info["convert_to_percent"]
    score_key = dataset_info["score_key"]
    has_invalid = dataset_info.get("has_invalid", True)  # 默认为True
    is_accuracy = dataset_info.get("is_accuracy", False)  # 默认为False
    parts = dataset_info.get("parts", {})
    control_groups = dataset_info.get("control_groups", [])
    
    # Print experimental group results
    for subset in subsets:
        performance_data = load_performance_data(model, subset, result_dir, wasr)
        if performance_data is None:
            print(f"    {subset}: [MISSING]")
            continue
            
        if subset in performance_data:
            subset_data = performance_data[subset]
            score = extract_score(subset_data, score_key)
            if score is not None:
                converted_score = convert_score(score, convert_to_percent, is_accuracy)
                
                # 根据是否有invalid字段决定打印哪些信息
                if has_invalid:
                    total = subset_data.get("total", "N/A") if isinstance(subset_data, dict) else "N/A"
                    invalid = subset_data.get("invalid", "N/A") if isinstance(subset_data, dict) else "N/A"
                    valid_count = total - invalid if isinstance(total, int) and isinstance(invalid, int) else "N/A"
                    print(f"    {subset}: {converted_score:.2f} (total: {total}, valid: {valid_count})")
                else:
                    # 对于GK、ML、DC，打印acc、valid、total和correct
                    acc = subset_data.get("acc", "N/A") if isinstance(subset_data, dict) else "N/A"
                    valid = subset_data.get("valid", "N/A") if isinstance(subset_data, dict) else "N/A"
                    total = subset_data.get("total", "N/A") if isinstance(subset_data, dict) else "N/A"
                    correct = subset_data.get("correct", "N/A") if isinstance(subset_data, dict) else "N/A"
                    print(f"    {subset}: {converted_score:.2f} (total: {total}, valid: {valid})")
                
                # Print subpart results
                if subset in parts:
                    for part in parts[subset]:
                        part_key = find_part_key(performance_data, part)
                        if part_key and part_key in performance_data:
                            part_data = performance_data[part_key]
                            part_score = extract_score(part_data, score_key)
                            if part_score is not None:
                                # 对于部分分数，也使用相同的转换逻辑
                                converted_part_score = convert_score(part_score, convert_to_percent, is_accuracy)
                                print(f"      {part}: {converted_part_score:.2f}")
                            else:
                                print(f"      {part}: [INVALID DATA]")
            else:
                print(f"    {subset}: [INVALID DATA]")
        else:
            print(f"    {subset}: [NO SCORE DATA]")
    
    # Print control group results
    if control_groups:
        print("    --- Control Groups ---")
        for control_group in control_groups:
            performance_data = load_performance_data(model, control_group, result_dir, wasr)
            if performance_data is None:
                print(f"    {control_group}: [MISSING]")
                continue
                
            if control_group in performance_data:
                control_data = performance_data[control_group]
                score = extract_score(control_data, score_key)
                if score is not None:
                    converted_score = convert_score(score, convert_to_percent, is_accuracy)
                    
                    # 根据是否有invalid字段决定打印哪些信息
                    if has_invalid:
                        total = control_data.get("total", "N/A") if isinstance(control_data, dict) else "N/A"
                        invalid = control_data.get("invalid", "N/A") if isinstance(control_data, dict) else "N/A"
                        valid_count = total - invalid if isinstance(total, int) and isinstance(invalid, int) else "N/A"
                        print(f"    {control_group}: {converted_score:.2f} (total: {total}, invalid: {invalid}, valid: {valid_count})")
                    else:
                        # 对于GK、ML、DC，打印acc、valid、total和correct
                        acc = control_data.get("acc", "N/A") if isinstance(control_data, dict) else "N/A"
                        valid = control_data.get("valid", "N/A") if isinstance(control_data, dict) else "N/A"
                        total = control_data.get("total", "N/A") if isinstance(control_data, dict) else "N/A"
                        correct = control_data.get("correct", "N/A") if isinstance(control_data, dict) else "N/A"
                        print(f"    {control_group}: {converted_score:.2f} (valid: {valid}, total: {total})")
                    
                    # Print control group subpart results
                    base_name = control_group.replace('_cmp', '')
                    if base_name in parts:
                        for part in parts[base_name]:
                            part_key = find_part_key(performance_data, part)
                            if part_key and part_key in performance_data:
                                part_data = performance_data[part_key]
                                part_score = extract_score(part_data, score_key)
                                if part_score is not None:
                                    # 对于部分分数，也使用相同的转换逻辑
                                    converted_part_score = convert_score(part_score, convert_to_percent, is_accuracy)
                                    print(f"      {part}: {converted_part_score:.2f}")
                                else:
                                    print(f"      {part}: [INVALID DATA]")

def find_part_key(performance_data, part_name):
    """Find the key corresponding to the part in performance data"""
    for key in performance_data.keys():
        if part_name == "logic":
           return "lr_logic"
        if part_name.lower() in key.lower() or key.lower().endswith(part_name.lower()):
            return key
    return None

def get_merge_rules(header_style):
    """定义所有表头合并规则（XlsxWriter格式，行/列从0开始）"""
    merge_rules = [
        # (0, 0, 0, 0, "Model", header_style),
        (0, 0, 2, 0, "Model/Dataset", header_style),
        # -------------------------- 第0行（顶层数据集名称，跨列） --------------------------
        # TIF：列1-8（共8列）
        (0, 1, 0, 8, "TIF(Text Instruction Following)", header_style),
        # TIF-En：列9-16（共8列）
        (0, 9, 0, 16, "TIF-En(English Text Instruction Following)", header_style),
        # SIF：列17-23（共7列）
        (0, 17, 0, 23, "SIF(Speech Instruction Following)", header_style),
        # SIF-En：列24-30（共7列）
        (0, 24, 0, 30, "SIF-En(English Speech Instruction Following)", header_style),
        # MTD：列31-34（共4列）
        (0, 31, 0, 34, "MTD(Multi-turn Dialogue)", header_style),
        # GK：列35-47（共13列）
        (0, 35, 0, 47, "GK(General Knowledge)", header_style),
        # ML：列48-55（共8列）
        (0, 48, 0, 55, "ML(Mathematical Logic)", header_style),
        # DC：列56-59（共4列）
        (0, 56, 0, 59, "DC(Discourse Comprehension)", header_style),
        # SV：列60-72（共13列）
        (0, 60, 0, 72, "SV(Speaker Variations)", header_style),
        # EV：列73-83（共11列）
        (0, 73, 0, 83, "EV(Environmental Variations)", header_style),
        # CV：列84-89（共6列）
        (0, 84, 0, 89, "CV(Content Variations)", header_style),

        # -------------------------- 第1行（子类别，跨列+跨行） --------------------------
        # TIF子类别（跨2行：第1-2行）
        (1, 1, 2, 1, "Avg.", header_style),
        (1, 2, 2, 2, "continuation", header_style),
        (1, 3, 2, 3, "creation", header_style),
        (1, 4, 2, 4, "empathy", header_style),
        (1, 5, 2, 5, "recommendation", header_style),
        (1, 6, 2, 6, "rewriting", header_style),
        (1, 7, 2, 7, "safety", header_style),
        (1, 8, 2, 8, "simulation", header_style),
        
        # TIF-En子类别（跨2行）
        (1, 9, 2, 9, "Avg.", header_style),
        (1, 10, 2, 10, "continuation_en", header_style),
        (1, 11, 2, 11, "creation_en", header_style),
        (1, 12, 2, 12, "empathy_en", header_style),
        (1, 13, 2, 13, "recommendation_en", header_style),
        (1, 14, 2, 14, "rewriting_en", header_style),
        (1, 15, 2, 15, "safety_en", header_style),
        (1, 16, 2, 16, "simulation_en", header_style),
        
        # SIF子类别（跨2行）
        (1, 17, 2, 17, "Avg.", header_style),
        (1, 18, 2, 18, "emotional_control", header_style),
        (1, 19, 2, 19, "language_control", header_style),
        (1, 20, 2, 20, "non_verbal_vocalization", header_style),
        (1, 21, 2, 21, "pacing_control", header_style),
        (1, 22, 2, 22, "style_control", header_style),
        (1, 23, 2, 23, "volume_control", header_style),
        
        # SIF-En子类别（跨2行）
        (1, 24, 2, 24, "Avg.", header_style),
        (1, 25, 2, 25, "emotional_control_en", header_style),
        (1, 26, 2, 26, "language_control_en", header_style),
        (1, 27, 2, 27, "non_verbal_vocalization_en", header_style),
        (1, 28, 2, 28, "pacing_control_en", header_style),
        (1, 29, 2, 29, "style_control_en", header_style),
        (1, 30, 2, 30, "volume_control_en", header_style),
        
        # MTD子类别（跨2行）
        (1, 31, 2, 31, "Avg.", header_style),
        (1, 32, 2, 32, "progression", header_style),
        (1, 33, 2, 33, "backtracking", header_style),
        (1, 34, 2, 34, "transition", header_style),
        
        # GK子类别（跨2行）
        (1, 35, 2, 35, "Avg.", header_style),
        (1, 36, 2, 36, "mathematics", header_style),
        (1, 37, 2, 37, "geography", header_style),
        (1, 38, 2, 38, "politics", header_style),
        (1, 39, 2, 39, "chemistry", header_style),
        (1, 40, 2, 40, "biology", header_style),
        (1, 41, 2, 41, "law", header_style),
        (1, 42, 2, 42, "physics", header_style),
        (1, 43, 2, 43, "history", header_style),
        (1, 44, 2, 44, "medicine", header_style),
        (1, 45, 2, 45, "economics", header_style),
        (1, 46, 2, 46, "sports", header_style),
        (1, 47, 2, 47, "culture", header_style),
        
        # ML子类别（特殊：logical_reasoning跨4列，不跨行）
        (1, 48, 2, 48, "Avg.", header_style),
        (1, 49, 2, 49, "basic_math", header_style),
        (1, 50, 2, 50, "math", header_style),
        (1, 51, 1, 55, "logical_reasoning", header_style),  # 仅第1行跨5列
        # (2, 51, 2, 51, "Avg.", header_style),          # 第2行子部分
        # (2, 52, 2, 52, "analysis", header_style),          # 第2行子部分
        # (2, 53, 2, 53, "induction", header_style),
        # (2, 54, 2, 54, "analogy", header_style),
        # (2, 55, 2, 55, "logic", header_style),
        
        # DC子类别（跨2行）
        (1, 56, 2, 56, "Avg.", header_style),
        (1, 57, 2, 57, "inference", header_style),
        (1, 58, 2, 58, "induction", header_style),
        (1, 59, 2, 59, "analysis", header_style),
        
        # SV子类别（特殊：age/accent/volume跨列，不跨行）
        (1, 60, 2, 60, "Avg.", header_style),
        (1, 61, 1, 63, "age", header_style),                # 第1行跨3列
        # (2, 61, 2, 61, "Avg.", header_style),              # 第2行子部分
        # (2, 62, 2, 62, "child", header_style),              # 第2行子部分
        # (2, 63, 2, 63, "elder", header_style),
        (1, 64, 1, 68, "accent", header_style),             # 第1行跨5列
        # (2, 64, 2, 64, "Avg.", header_style),
        # (2, 65, 2, 65, "tianjin", header_style),
        # (2, 66, 2, 66, "beijing", header_style),
        # (2, 67, 2, 67, "dongbei", header_style),
        # (2, 68, 2, 68, "sichuan", header_style),
        (1, 69, 1, 71, "volume", header_style),             # 第1行跨3列
        # (2, 69, 2, 69, "Avg.", header_style),
        # (2, 70, 2, 70, "down", header_style),
        # (2, 71, 2, 71, "up", header_style),
        # (1, 72, 1, 72, "speed", header_style),              # 跨1行
        # (2, 72, 2, 72, "Avg.", header_style),           
        
        # EV子类别（特殊：non_vocal/vocal跨列，不跨行）
        (1, 73, 2, 73, "Avg.", header_style),
        (1, 74, 1, 77, "non_vocal_noise", header_style),    # 第1行跨4列
        # (2, 74, 2, 74, "Avg.", header_style),
        # (2, 75, 2, 75, "echo", header_style),
        # (2, 76, 2, 76, "outdoors", header_style),
        # (2, 77, 2, 77, "farField", header_style),
        (1, 78, 1, 82, "vocal_noise", header_style),        # 第1行跨5列
        # (2, 78, 2, 78, "Avg.", header_style),
        # (2, 79, 2, 79, "tvPlayback", header_style),
        # (2, 80, 2, 80, "backgroundChat", header_style),
        # (2, 81, 2, 81, "vocalMusic", header_style),
        # (2, 82, 2, 82, "voiceAnnouncement", header_style),
        # (1, 83, 1, 83, "unstable_signal", header_style),    # 跨1行
        # (2, 83, 2, 83, "Avg.", header_style),
        
        # CV子类别（跨2行）
        (1, 84, 2, 84, "Avg.", header_style),
        (1, 85, 2, 85, "casual_talk", header_style),
        (1, 86, 2, 86, "mispronunciation", header_style),
        (1, 87, 2, 87, "grammatical_error", header_style),
        (1, 88, 2, 88, "topic_shift", header_style),
        (1, 89, 2, 89, "code_switching", header_style),

    ]
    return merge_rules

def get_single_cell(header_style):
    single_cells = [
        # (0, 0, "Model", header_style),
        (2, 51, "Avg.", header_style), 
        (2, 52, "analysis", header_style), 
        (2, 53, "induction", header_style),
        (2, 54, "analogy", header_style),
        (2, 55, "logic", header_style),
        (2, 61, "Avg.", header_style),
        (2, 62, "child", header_style),
        (2, 63, "elder", header_style),
        (2, 64, "Avg.", header_style),
        (2, 65, "tianjin", header_style),
        (2, 66, "beijing", header_style),
        (2, 67, "dongbei", header_style),
        (2, 68, "sichuan", header_style),
        (2, 69, "Avg.", header_style),
        (2, 70, "down", header_style),
        (2, 71, "up", header_style),
        (1, 72, "speed", header_style),
        (2, 72, "Avg.", header_style),
        (1, 73, "Avg.", header_style),
        (2, 74, "Avg.", header_style),
        (2, 75, "echo", header_style),
        (2, 76, "outdoors", header_style),
        (2, 77, "farField", header_style),
        (2, 78, "Avg.", header_style),
        (2, 79, "tvPlayback", header_style),
        (2, 80, "backgroundChat", header_style),
        (2, 81, "vocalMusic", header_style),
        (2, 82, "voiceAnnouncement", header_style),
        (1, 83, "unstable_signal", header_style),
        (2, 83, "Avg.", header_style),
    ]
    return single_cells

def create_summary_sheet(workbook, header_style, data_style, model_name_style, result_dir, all_models_results, wasr=False):
    # 定义Summary的列（数据集大类）
    summary_columns = list(DATASET_STRUCTURE.keys())
    # 构建Summary数据
    summary_data = []
    for model in all_models_results.keys():
        row = {"Model": model}
        for dataset in summary_columns:
            score = all_models_results[model].get(dataset)
            row[dataset] = f"{score:.2f}" if score is not None else "N/A"
        summary_data.append(row)
    # 写入Excel
    summary_sheet = workbook.add_worksheet("Summary")
    # 写入表头
    summary_sheet.write(0, 0, "Model/Dataset", header_style)
    for col_idx, col_name in enumerate(summary_columns, 1):
        summary_sheet.write(0, col_idx, col_name, header_style)
    # 写入数据
    for row_idx, row in enumerate(summary_data, 1):
        summary_sheet.write(row_idx, 0, row["Model"], model_name_style)
        for col_idx, col_name in enumerate(summary_columns, 1):
            summary_sheet.write(row_idx, col_idx, row[col_name], data_style)
    # 调整列宽
    summary_sheet.set_column(0, 0, 25)
    for col in range(1, len(summary_columns) + 1):
        summary_sheet.set_column(col, col, 44)

def create_detail_sheet(workbook, header_style, data_style, model_name_style, result_dir, all_models_results, wasr=False):
    detailed_sheet = workbook.add_worksheet("Detailed Results")
    # 创建表头（前3行为表头）
    single_cells = get_single_cell(header_style)
    for row_idx, col_idx, value, style in single_cells:
        detailed_sheet.write(row_idx, col_idx, value, style)
    merge_rules = get_merge_rules(header_style)
    for rule in merge_rules:
        start_row, start_col, end_row, end_col, value, style = rule
        detailed_sheet.merge_range(start_row, start_col, end_row, end_col, value, style)
    # - 写入数据行（从第3行开始）
    current_row = 3  
    # 遍历所有模型
    for model, model_results in all_models_results.items():
        # 写入模型名称（第0列）
        detailed_sheet.write(current_row, 0, model, model_name_style)
        
        # 当前列索引（从第1列开始）
        current_col = 1
        
        # 1. 处理TIF数据
        tif_data = model_results.get("TIF(Text Instruction Following)")
        # 写入TIF平均值
        detailed_sheet.write(current_row, current_col, f"{tif_data:.2f}" if tif_data is not None else "N/A", data_style)
        current_col += 1
        # 写入TIF各子集数据
        for subset in ["continuation", "creation", "empathy", "recommendation", "rewriting", "safety", "simulation"]:
            performance_data = load_performance_data(model, subset, result_dir, wasr)
            score = None
            if performance_data and subset in performance_data:
                score = extract_score(performance_data[subset], DATASET_STRUCTURE["TIF(Text Instruction Following)"]["score_key"])
                if score is not None:
                    score = convert_score(score, True)  # TIF需要转为百分比
            detailed_sheet.write(current_row, current_col, f"{score:.2f}" if score is not None else "N/A", data_style)
            current_col += 1
        
        # 2. 处理TIF-En数据
        tif_en_data = model_results.get("TIF-En(English Text Instruction Following)")
        # 写入TIF-En平均值
        detailed_sheet.write(current_row, current_col, f"{tif_en_data:.2f}" if tif_en_data is not None else "N/A", data_style)
        current_col += 1
        # 写入TIF-En各子集数据
        for subset in ["continuation_en", "creation_en", "empathy_en", "recommendation_en", "rewriting_en", "safety_en", "simulation_en"]:
            performance_data = load_performance_data(model, subset, result_dir, wasr)
            score = None
            if performance_data and subset in performance_data:
                score = extract_score(performance_data[subset], DATASET_STRUCTURE["TIF-En(English Text Instruction Following)"]["score_key"])
                if score is not None:
                    score = convert_score(score, True)  # TIF-En需要转为百分比
            detailed_sheet.write(current_row, current_col, f"{score:.2f}" if score is not None else "N/A", data_style)
            current_col += 1
        
        # 3. 处理SIF数据
        sif_data = model_results.get("SIF(Speech Instruction Following)")
        # 写入SIF平均值
        detailed_sheet.write(current_row, current_col, f"{sif_data:.2f}" if sif_data is not None else "N/A", data_style)
        current_col += 1
        # 写入SIF各子集数据
        for subset in ["emotional_control", "language_control", "non_verbal_vocalization", "pacing_control", "style_control", "volume_control"]:
            performance_data = load_performance_data(model, subset, result_dir, wasr)
            score = None
            if performance_data and subset in performance_data:
                score = extract_score(performance_data[subset], DATASET_STRUCTURE["SIF(Speech Instruction Following)"]["score_key"])
                if score is not None:
                    score = convert_score(score, True)  # SIF需要转为百分比
            detailed_sheet.write(current_row, current_col, f"{score:.2f}" if score is not None else "N/A", data_style)
            current_col += 1
        
        # 4. 处理SIF-En数据
        sif_en_data = model_results.get("SIF-En(English Speech Instruction Following)")
        # 写入SIF-En平均值
        detailed_sheet.write(current_row, current_col, f"{sif_en_data:.2f}" if sif_en_data is not None else "N/A", data_style)
        current_col += 1
        # 写入SIF-En各子集数据
        for subset in ["emotional_control_en", "language_control_en", "non_verbal_vocalization_en", "pacing_control_en", "style_control_en", "volume_control_en"]:
            performance_data = load_performance_data(model, subset, result_dir, wasr)
            score = None
            if performance_data and subset in performance_data:
                score = extract_score(performance_data[subset], DATASET_STRUCTURE["SIF-En(English Speech Instruction Following)"]["score_key"])
                if score is not None:
                    score = convert_score(score, True)  # SIF-En需要转为百分比
            detailed_sheet.write(current_row, current_col, f"{score:.2f}" if score is not None else "N/A", data_style)
            current_col += 1
        
        # 5. 处理MTD数据
        mtd_data = model_results.get("MTD(Multi-turn Dialogue)")
        # 写入MTD平均值
        detailed_sheet.write(current_row, current_col, f"{mtd_data:.2f}" if mtd_data is not None else "N/A", data_style)
        current_col += 1
        # 写入MTD各子集数据
        for subset in ["progression", "backtracking", "transition"]:
            performance_data = load_performance_data(model, subset, result_dir, wasr)
            score = None
            if performance_data and subset in performance_data:
                score = extract_score(performance_data[subset], DATASET_STRUCTURE["MTD(Multi-turn Dialogue)"]["score_key"])
                if score is not None:
                    score = convert_score(score, True)  # MTD需要转为百分比
            detailed_sheet.write(current_row, current_col, f"{score:.2f}" if score is not None else "N/A", data_style)
            current_col += 1
        
        # 6. 处理GK数据
        gk_data = model_results.get("GK(General Knowledge)")
        # 写入GK平均值
        detailed_sheet.write(current_row, current_col, f"{gk_data:.2f}" if gk_data is not None else "N/A", data_style)
        current_col += 1
        # 写入GK各部分数据
        for part in ["mathematics", "geography", "politics", "chemistry", "biology", "law", "physics", "history", "medicine", "economics", "sports", "culture"]:
            performance_data = load_performance_data(model, "general_knowledge", result_dir, wasr)
            score = None
            if performance_data:
                part_key = find_part_key(performance_data, part)
                if part_key in performance_data:
                    score = extract_score(performance_data[part_key], DATASET_STRUCTURE["GK(General Knowledge)"]["score_key"])
                    if score is not None:
                        score = convert_score(score, True, is_accuracy=True)  # GK是准确率，需要转为百分比
            detailed_sheet.write(current_row, current_col, f"{score:.2f}" if score is not None else "N/A", data_style)
            current_col += 1
        
        # 7. 处理ML数据
        ml_data = model_results.get("ML(Mathematical Logic)")
        # 写入ML平均值
        detailed_sheet.write(current_row, current_col, f"{ml_data:.2f}" if ml_data is not None else "N/A", data_style)
        current_col += 1
        # 写入basic_math和math
        for subset in ["basic_math", "math"]:
            performance_data = load_performance_data(model, subset, result_dir, wasr)
            score = None
            if performance_data and subset in performance_data:
                score = extract_score(performance_data[subset], DATASET_STRUCTURE["ML(Mathematical Logic)"]["score_key"])
                if score is not None:
                    score = convert_score(score, True, is_accuracy=True)  # ML是准确率，需要转为百分比
            detailed_sheet.write(current_row, current_col, f"{score:.2f}" if score is not None else "N/A", data_style)
            current_col += 1
        # 写入logical_reasoning及其各部分
        performance_data = load_performance_data(model, "logical_reasoning", result_dir, wasr)
        # logical_reasoning平均值
        lr_score = None
        if performance_data and "logical_reasoning" in performance_data:
            lr_score = extract_score(performance_data["logical_reasoning"], DATASET_STRUCTURE["ML(Mathematical Logic)"]["score_key"])
            if lr_score is not None:
                lr_score = convert_score(lr_score, True, is_accuracy=True)
        detailed_sheet.write(current_row, current_col, f"{lr_score:.2f}" if lr_score is not None else "N/A", data_style)
        current_col += 1
        # logical_reasoning各部分
        for part in ["analysis", "induction", "analogy", "logic"]:
            score = None
            if performance_data:
                part_key = find_part_key(performance_data, part)
                if part_key in performance_data:
                    score = extract_score(performance_data[part_key], DATASET_STRUCTURE["ML(Mathematical Logic)"]["score_key"])
                    if score is not None:
                        score = convert_score(score, True, is_accuracy=True)
            detailed_sheet.write(current_row, current_col, f"{score:.2f}" if score is not None else "N/A", data_style)
            current_col += 1
        
        # 8. 处理DC数据
        dc_data = model_results.get("DC(Discourse Comprehension)")
        # 写入DC平均值
        detailed_sheet.write(current_row, current_col, f"{dc_data:.2f}" if dc_data is not None else "N/A", data_style)
        current_col += 1
        # 写入DC各部分数据
        for part in ["inference", "induction", "analysis"]:
            performance_data = load_performance_data(model, "discourse_comprehension", result_dir, wasr)
            score = None
            if performance_data:
                part_key = find_part_key(performance_data, part)
                if part_key in performance_data:
                    score = extract_score(performance_data[part_key], DATASET_STRUCTURE["DC(Discourse Comprehension)"]["score_key"])
                    if score is not None:
                        score = convert_score(score, True, is_accuracy=True)  # DC是准确率，需要转为百分比
            detailed_sheet.write(current_row, current_col, f"{score:.2f}" if score is not None else "N/A", data_style)
            current_col += 1
        
        # 9. 处理SV数据（格式为Exp/Cmp）
        _, sv_avg, sv_cmp_avg = calculate_dataset_average(model, "SV(Speaker Variations)", DATASET_STRUCTURE["SV(Speaker Variations)"], result_dir, wasr)
        # 写入SV平均值
        sv_avg_str = f"{sv_avg:.2f} / {sv_cmp_avg:.2f}" if sv_avg is not None and sv_cmp_avg is not None else "N/A"
        detailed_sheet.write(current_row, current_col, sv_avg_str, data_style)
        current_col += 1
        
        # 处理age及其部分
        # age平均值
        age_score = None
        age_cmp_score = None
        age_perf = load_performance_data(model, "age", result_dir, wasr)
        if age_perf and "age" in age_perf:
            age_score = extract_score(age_perf["age"], DATASET_STRUCTURE["SV(Speaker Variations)"]["score_key"])
            if age_score is not None:
                age_score = convert_score(age_score, True)
        age_cmp_perf = load_performance_data(model, "age_cmp", result_dir, wasr)
        if age_cmp_perf and "age_cmp" in age_cmp_perf:
            age_cmp_score = extract_score(age_cmp_perf["age_cmp"], DATASET_STRUCTURE["SV(Speaker Variations)"]["score_key"])
            if age_cmp_score is not None:
                age_cmp_score = convert_score(age_cmp_score, True)
        age_avg_str = f"{age_score:.2f} / {age_cmp_score:.2f}" if age_score is not None and age_cmp_score is not None else "N/A"
        detailed_sheet.write(current_row, current_col, age_avg_str, data_style)
        current_col += 1
        # age各部分
        for part in ["child", "elder"]:
            exp_score = None
            cmp_score = None
            # 实验组分数
            if age_perf:
                part_key = find_part_key(age_perf, part)
                if part_key in age_perf:
                    exp_score = extract_score(age_perf[part_key], DATASET_STRUCTURE["SV(Speaker Variations)"]["score_key"])
                    if exp_score is not None:
                        exp_score = convert_score(exp_score, True)
            # 对照组分数
            if age_cmp_perf:
                part_key = find_part_key(age_cmp_perf, part)
                if part_key in age_cmp_perf:
                    cmp_score = extract_score(age_cmp_perf[part_key], DATASET_STRUCTURE["SV(Speaker Variations)"]["score_key"])
                    if cmp_score is not None:
                        cmp_score = convert_score(cmp_score, True)
            part_str = f"{exp_score:.2f} / {cmp_score:.2f}" if exp_score is not None and cmp_score is not None else "N/A"
            detailed_sheet.write(current_row, current_col, part_str, data_style)
            current_col += 1
        
        # 处理accent及其部分
        # accent平均值
        accent_score = None
        accent_cmp_score = None
        accent_perf = load_performance_data(model, "accent", result_dir, wasr)
        if accent_perf and "accent" in accent_perf:
            accent_score = extract_score(accent_perf["accent"], DATASET_STRUCTURE["SV(Speaker Variations)"]["score_key"])
            if accent_score is not None:
                accent_score = convert_score(accent_score, True)
        accent_cmp_perf = load_performance_data(model, "accent_cmp", result_dir, wasr)
        if accent_cmp_perf and "accent_cmp" in accent_cmp_perf:
            accent_cmp_score = extract_score(accent_cmp_perf["accent_cmp"], DATASET_STRUCTURE["SV(Speaker Variations)"]["score_key"])
            if accent_cmp_score is not None:
                accent_cmp_score = convert_score(accent_cmp_score, True)
        accent_avg_str = f"{accent_score:.2f} / {accent_cmp_score:.2f}" if accent_score is not None and accent_cmp_score is not None else "N/A"
        detailed_sheet.write(current_row, current_col, accent_avg_str, data_style)
        current_col += 1
        # accent各部分
        for part in ["tianjin", "beijing", "dongbei", "sichuan"]:
            exp_score = None
            cmp_score = None
            # 实验组分数
            if accent_perf:
                part_key = find_part_key(accent_perf, part)
                if part_key in accent_perf:
                    exp_score = extract_score(accent_perf[part_key], DATASET_STRUCTURE["SV(Speaker Variations)"]["score_key"])
                    if exp_score is not None:
                        exp_score = convert_score(exp_score, True)
            # 对照组分数
            if accent_cmp_perf:
                part_key = find_part_key(accent_cmp_perf, part)
                if part_key in accent_cmp_perf:
                    cmp_score = extract_score(accent_cmp_perf[part_key], DATASET_STRUCTURE["SV(Speaker Variations)"]["score_key"])
                    if cmp_score is not None:
                        cmp_score = convert_score(cmp_score, True)
            part_str = f"{exp_score:.2f} / {cmp_score:.2f}" if exp_score is not None and cmp_score is not None else "N/A"
            detailed_sheet.write(current_row, current_col, part_str, data_style)
            current_col += 1
        
        # 处理volume及其部分
        # volume平均值
        volume_score = None
        volume_cmp_score = None
        volume_perf = load_performance_data(model, "volume", result_dir, wasr)
        if volume_perf and "volume" in volume_perf:
            volume_score = extract_score(volume_perf["volume"], DATASET_STRUCTURE["SV(Speaker Variations)"]["score_key"])
            if volume_score is not None:
                volume_score = convert_score(volume_score, True)
        volume_cmp_perf = load_performance_data(model, "volume_cmp", result_dir, wasr)
        if volume_cmp_perf and "volume_cmp" in volume_cmp_perf:
            volume_cmp_score = extract_score(volume_cmp_perf["volume_cmp"], DATASET_STRUCTURE["SV(Speaker Variations)"]["score_key"])
            if volume_cmp_score is not None:
                volume_cmp_score = convert_score(volume_cmp_score, True)
        volume_avg_str = f"{volume_score:.2f} / {volume_cmp_score:.2f}" if volume_score is not None and volume_cmp_score is not None else "N/A"
        detailed_sheet.write(current_row, current_col, volume_avg_str, data_style)
        current_col += 1
        # volume各部分
        for part in ["down", "up"]:
            exp_score = None
            cmp_score = None
            # 实验组分数
            if volume_perf:
                part_key = find_part_key(volume_perf, part)
                if part_key in volume_perf:
                    exp_score = extract_score(volume_perf[part_key], DATASET_STRUCTURE["SV(Speaker Variations)"]["score_key"])
                    if exp_score is not None:
                        exp_score = convert_score(exp_score, True)
            # 对照组分数
            if volume_cmp_perf:
                part_key = find_part_key(volume_cmp_perf, part)
                if part_key in volume_cmp_perf:
                    cmp_score = extract_score(volume_cmp_perf[part_key], DATASET_STRUCTURE["SV(Speaker Variations)"]["score_key"])
                    if cmp_score is not None:
                        cmp_score = convert_score(cmp_score, True)
            part_str = f"{exp_score:.2f} / {cmp_score:.2f}" if exp_score is not None and cmp_score is not None else "N/A"
            detailed_sheet.write(current_row, current_col, part_str, data_style)
            current_col += 1
        
        # 处理speed
        speed_score = None
        speed_cmp_score = None
        speed_perf = load_performance_data(model, "speed", result_dir, wasr)
        if speed_perf and "speed" in speed_perf:
            speed_score = extract_score(speed_perf["speed"], DATASET_STRUCTURE["SV(Speaker Variations)"]["score_key"])
            if speed_score is not None:
                speed_score = convert_score(speed_score, True)
        speed_cmp_perf = load_performance_data(model, "speed_cmp", result_dir, wasr)
        if speed_cmp_perf and "speed_cmp" in speed_cmp_perf:
            speed_cmp_score = extract_score(speed_cmp_perf["speed_cmp"], DATASET_STRUCTURE["SV(Speaker Variations)"]["score_key"])
            if speed_cmp_score is not None:
                speed_cmp_score = convert_score(speed_cmp_score, True)
        speed_str = f"{speed_score:.2f} / {speed_cmp_score:.2f}" if speed_score is not None and speed_cmp_score is not None else "N/A"
        detailed_sheet.write(current_row, current_col, speed_str, data_style)
        current_col += 1
        
        # 10. 处理EV数据（格式为Exp/Cmp）
        _, ev_avg, ev_cmp_avg = calculate_dataset_average(model, "EV(Environmental Variations)", DATASET_STRUCTURE["EV(Environmental Variations)"], result_dir, wasr)
        # 写入EV平均值
        ev_avg_str = f"{ev_avg:.2f} / {ev_cmp_avg:.2f}" if ev_avg is not None and ev_cmp_avg is not None else "N/A"
        detailed_sheet.write(current_row, current_col, ev_avg_str, data_style)
        current_col += 1
        
        # 处理non_vocal_noise及其部分
        # non_vocal_noise平均值
        nv_score = None
        nv_cmp_score = None
        nv_perf = load_performance_data(model, "non_vocal_noise", result_dir, wasr)
        if nv_perf and "non_vocal_noise" in nv_perf:
            nv_score = extract_score(nv_perf["non_vocal_noise"], DATASET_STRUCTURE["EV(Environmental Variations)"]["score_key"])
            if nv_score is not None:
                nv_score = convert_score(nv_score, True)
        nv_cmp_perf = load_performance_data(model, "non_vocal_noise_cmp", result_dir, wasr)
        if nv_cmp_perf and "non_vocal_noise_cmp" in nv_cmp_perf:
            nv_cmp_score = extract_score(nv_cmp_perf["non_vocal_noise_cmp"], DATASET_STRUCTURE["EV(Environmental Variations)"]["score_key"])
            if nv_cmp_score is not None:
                nv_cmp_score = convert_score(nv_cmp_score, True)
        nv_avg_str = f"{nv_score:.2f} / {nv_cmp_score:.2f}" if nv_score is not None and nv_cmp_score is not None else "N/A"
        detailed_sheet.write(current_row, current_col, nv_avg_str, data_style)
        current_col += 1
        # non_vocal_noise各部分
        for part in ["echo", "outdoors", "farField"]:
            exp_score = None
            cmp_score = None
            # 实验组分数
            if nv_perf:
                part_key = find_part_key(nv_perf, part)
                if part_key in nv_perf:
                    exp_score = extract_score(nv_perf[part_key], DATASET_STRUCTURE["EV(Environmental Variations)"]["score_key"])
                    if exp_score is not None:
                        exp_score = convert_score(exp_score, True)
            # 对照组分数
            if nv_cmp_perf:
                part_key = find_part_key(nv_cmp_perf, part)
                if part_key in nv_cmp_perf:
                    cmp_score = extract_score(nv_cmp_perf[part_key], DATASET_STRUCTURE["EV(Environmental Variations)"]["score_key"])
                    if cmp_score is not None:
                        cmp_score = convert_score(cmp_score, True)
            part_str = f"{exp_score:.2f} / {cmp_score:.2f}" if exp_score is not None and cmp_score is not None else "N/A"
            detailed_sheet.write(current_row, current_col, part_str, data_style)
            current_col += 1
        
        # 处理vocal_noise及其部分
        # vocal_noise平均值
        v_score = None
        v_cmp_score = None
        v_perf = load_performance_data(model, "vocal_noise", result_dir, wasr)
        if v_perf and "vocal_noise" in v_perf:
            v_score = extract_score(v_perf["vocal_noise"], DATASET_STRUCTURE["EV(Environmental Variations)"]["score_key"])
            if v_score is not None:
                v_score = convert_score(v_score, True)
        v_cmp_perf = load_performance_data(model, "vocal_noise_cmp", result_dir, wasr)
        if v_cmp_perf and "vocal_noise_cmp" in v_cmp_perf:
            v_cmp_score = extract_score(v_cmp_perf["vocal_noise_cmp"], DATASET_STRUCTURE["EV(Environmental Variations)"]["score_key"])
            if v_cmp_score is not None:
                v_cmp_score = convert_score(v_cmp_score, True)
        v_avg_str = f"{v_score:.2f} / {v_cmp_score:.2f}" if v_score is not None and v_cmp_score is not None else "N/A"
        detailed_sheet.write(current_row, current_col, v_avg_str, data_style)
        current_col += 1
        # vocal_noise各部分
        for part in ["tvPlayback", "backgroundChat", "vocalMusic", "voiceAnnouncement"]:
            exp_score = None
            cmp_score = None
            # 实验组分数
            if v_perf:
                part_key = find_part_key(v_perf, part)
                if part_key in v_perf:
                    exp_score = extract_score(v_perf[part_key], DATASET_STRUCTURE["EV(Environmental Variations)"]["score_key"])
                    if exp_score is not None:
                        exp_score = convert_score(exp_score, True)
            # 对照组分数
            if v_cmp_perf:
                part_key = find_part_key(v_cmp_perf, part)
                if part_key in v_cmp_perf:
                    cmp_score = extract_score(v_cmp_perf[part_key], DATASET_STRUCTURE["EV(Environmental Variations)"]["score_key"])
                    if cmp_score is not None:
                        cmp_score = convert_score(cmp_score, True)
            part_str = f"{exp_score:.2f} / {cmp_score:.2f}" if exp_score is not None and cmp_score is not None else "N/A"
            detailed_sheet.write(current_row, current_col, part_str, data_style)
            current_col += 1
        
        # 处理unstable_signal
        us_score = None
        us_cmp_score = None
        us_perf = load_performance_data(model, "unstable_signal", result_dir, wasr)
        if us_perf and "unstable_signal" in us_perf:
            us_score = extract_score(us_perf["unstable_signal"], DATASET_STRUCTURE["EV(Environmental Variations)"]["score_key"])
            if us_score is not None:
                us_score = convert_score(us_score, True)
        us_cmp_perf = load_performance_data(model, "unstable_signal_cmp", result_dir, wasr)
        if us_cmp_perf and "unstable_signal_cmp" in us_cmp_perf:
            us_cmp_score = extract_score(us_cmp_perf["unstable_signal_cmp"], DATASET_STRUCTURE["EV(Environmental Variations)"]["score_key"])
            if us_cmp_score is not None:
                us_cmp_score = convert_score(us_cmp_score, True)
        us_str = f"{us_score:.2f} / {us_cmp_score:.2f}" if us_score is not None and us_cmp_score is not None else "N/A"
        detailed_sheet.write(current_row, current_col, us_str, data_style)
        current_col += 1
        
        # 11. 处理CV数据（格式为Exp/Cmp）
        _, cv_avg, cv_cmp_avg = calculate_dataset_average(model, "CV(Content Variations)", DATASET_STRUCTURE["CV(Content Variations)"], result_dir, wasr)
        # 写入CV平均值
        cv_avg_str = f"{cv_avg:.2f} / {cv_cmp_avg:.2f}" if cv_avg is not None and cv_cmp_avg is not None else "N/A"
        detailed_sheet.write(current_row, current_col, cv_avg_str, data_style)
        current_col += 1
        
        # 写入CV各子集数据
        for subset in ["casual_talk", "mispronunciation", "grammatical_error", "topic_shift", "code_switching"]:
            # 实验组分数
            exp_score = None
            exp_perf = load_performance_data(model, subset, result_dir, wasr)
            if exp_perf and subset in exp_perf:
                exp_score = extract_score(exp_perf[subset], DATASET_STRUCTURE["CV(Content Variations)"]["score_key"])
                if exp_score is not None:
                    exp_score = convert_score(exp_score, True)
            
            # 对照组分数
            cmp_score = None
            cmp_subset = f"{subset}_cmp"
            cmp_perf = load_performance_data(model, cmp_subset, result_dir, wasr)
            if cmp_perf and cmp_subset in cmp_perf:
                cmp_score = extract_score(cmp_perf[cmp_subset], DATASET_STRUCTURE["CV(Content Variations)"]["score_key"])
                if cmp_score is not None:
                    cmp_score = convert_score(cmp_score, True)
            
            # 写入合并分数
            subset_str = f"{exp_score:.2f} / {cmp_score:.2f}" if exp_score is not None and cmp_score is not None else "N/A"
            detailed_sheet.write(current_row, current_col, subset_str, data_style)
            current_col += 1
        
        # 移动到下一行
        current_row += 1

    # 4. 调整列宽
    detailed_sheet.set_column(0, 0, 25)  # Model列
    for col in range(1, 90):
        detailed_sheet.set_column(col, col, 20)  # 其他列

def export_to_excel(all_models_results, result_dir, output_file="evaluation_results.xlsx", wasr=False):
     # 1. 创建XlsxWriter工作簿
    workbook = xlsxwriter.Workbook(output_file)
    # 定义样式
    header_style = workbook.add_format({
        "bold": True,
        "bg_color": "#DDDDDD",
        "align": "center",
        "valign": "vcenter",
        "border": 1  # 细边框
    })
    data_style = workbook.add_format({
        "align": "center",
        "valign": "vcenter",
        "border": 1
    })
    model_name_style = workbook.add_format({
        "bold": True,
        "align": "left",
        "valign": "vcenter",
        "border": 1
    })
    # 2.1 创建Summary Results工作表
    create_summary_sheet(workbook, header_style, data_style, model_name_style, result_dir, all_models_results, wasr)
    # 2.2 创建Detailed Results工作表
    create_detail_sheet(workbook, header_style, data_style, model_name_style, result_dir, all_models_results, wasr)
    workbook.close()
    print(f"Excel文件已生成：{output_file}")

def main():
    parser = argparse.ArgumentParser(description='Statistics of model evaluation results')
    parser.add_argument('--model', nargs='+', required=True, help='Target model name list')
    parser.add_argument('--result_dir', default='./eval_result', help='Directory containing evaluation results')
    parser.add_argument('--export_excel', action='store_true', help='Export results to Excel file')
    parser.add_argument('--output_file', default='evaluation_results.xlsx', help='Output Excel file name')
    parser.add_argument('--wasr', action='store_true', help='Audio to audio with asr')
    
    args = parser.parse_args()
    
    # 用于存储所有模型的结果
    all_models_results = {}
    
    for model in args.model:
        print(f"\n{'='*60}")
        print(f"Model: {model}")
        print(f"{'='*60}")
        
        model_results = {}
        
        for dataset_name, dataset_info in DATASET_STRUCTURE.items():
            print(f"\n{dataset_name}:")
            
            # Calculate dataset average score
            all_exist, avg_score, avg_control_score = calculate_dataset_average(model, dataset_name, dataset_info, args.result_dir, args.wasr)
            
            if not all_exist:
                print("  [INCOMPLETE - Results incomplete]")
                model_results[dataset_name] = None
            else:
                if avg_score is None:
                    print("  [Cannot calculate average - All subsets have 0 valid data]")
                    model_results[dataset_name] = None
                elif avg_control_score is not None:
                    print(f"  Weighted Average (Exp/Cmp): {avg_score:.2f}/{avg_control_score:.2f}")
                    model_results[dataset_name] = avg_score
                else:
                    print(f"  Weighted Average: {avg_score:.2f}")
                    model_results[dataset_name] = avg_score
            
            # Print subset detailed results
            print_subset_results(model, dataset_name, dataset_info, args.result_dir, args.wasr)
        
        # 存储当前模型的结果
        all_models_results[model] = model_results
    
    # 如果指定了导出Excel，则调用导出函数
    if args.export_excel:
        export_to_excel(all_models_results, args.result_dir, args.output_file, args.wasr)

if __name__ == "__main__":
    main()

